"""``XlsxFormat`` — Microsoft Excel ``.xlsx`` (Office Open XML) encoder/decoder.

Reads use ``openpyxl`` in ``read_only=True`` / ``data_only=True`` mode:
that disables macro execution and surfaces evaluated cell values rather
than formulas. Writes use ``xlsxwriter``, which is the faster path for
producing fresh workbooks.

XLSX is a zipped XML bundle: random access requires the full archive,
so this is a :class:`BatchFileFormat`.

Security: pirn does not sandbox openpyxl. Macros (vbaProject.bin) are
ignored by ``read_only`` mode, but malformed archives may still trigger
upstream library bugs. Treat untrusted ``.xlsx`` payloads accordingly.

Install: ``pip install pirn[xlsx]``.
"""

from __future__ import annotations

import io
from collections.abc import Iterable, Mapping, Sequence
from typing import Any

from pirn.domains.connectors.file_formats.batch_file_format import (
    BatchFileFormat,
)


class XlsxFormat(BatchFileFormat):
    """Whole-file XLSX encoder/decoder."""

    def __init__(
        self,
        sheet_name: str = "Sheet1",
        has_header: bool = True,
        column_names: Sequence[str] | None = None,
    ) -> None:
        if not isinstance(sheet_name, str) or not sheet_name:
            raise ValueError("XlsxFormat: sheet_name must be a non-empty string")
        if not isinstance(has_header, bool):
            raise TypeError(
                f"XlsxFormat: has_header must be a bool, got {type(has_header).__name__}"
            )
        if column_names is not None:
            if not isinstance(column_names, Sequence) or isinstance(column_names, (str, bytes)):
                raise TypeError(
                    "XlsxFormat: column_names must be a sequence of "
                    f"strings, got {type(column_names).__name__}"
                )
            for col in column_names:
                if not isinstance(col, str) or not col:
                    raise ValueError(
                        f"XlsxFormat: every column name must be a non-empty string, got {col!r}"
                    )
        if not has_header and column_names is None:
            raise ValueError("XlsxFormat: column_names is required when has_header=False")
        self._sheet_name = sheet_name
        self._has_header = has_header
        self._column_names: tuple[str, ...] | None = (
            tuple(column_names) if column_names is not None else None
        )

    @property
    def name(self) -> str:
        return "xlsx"

    @property
    def sheet_name(self) -> str:
        return self._sheet_name

    @property
    def has_header(self) -> bool:
        return self._has_header

    @property
    def column_names(self) -> tuple[str, ...] | None:
        return self._column_names

    async def _decode_full(self, payload: bytes) -> Iterable[Mapping[str, Any]]:
        openpyxl = self._load_openpyxl()
        workbook = openpyxl.load_workbook(
            io.BytesIO(payload),
            read_only=True,
            data_only=True,
        )
        try:
            if self._sheet_name not in workbook.sheetnames:
                raise ValueError(
                    "XlsxFormat: sheet "
                    f"{self._sheet_name!r} not found in workbook; "
                    f"available sheets: {workbook.sheetnames}"
                )
            worksheet = workbook[self._sheet_name]
            row_iter = worksheet.iter_rows(values_only=True)
            columns: tuple[str, ...]
            if self._has_header:
                try:
                    header_row = next(row_iter)
                except StopIteration:
                    return []
                if self._column_names is not None:
                    columns = self._column_names
                else:
                    columns = tuple(str(cell) if cell is not None else "" for cell in header_row)
            else:
                if self._column_names is None:
                    raise RuntimeError(
                        "XlsxFormat: missing column_names while "
                        "has_header=False (should be unreachable)"
                    )
                columns = self._column_names
            records: list[Mapping[str, Any]] = []
            for row in row_iter:
                if all(cell is None for cell in row):
                    continue
                record: dict[str, Any] = {}
                for index, column in enumerate(columns):
                    record[column] = row[index] if index < len(row) else None
                records.append(record)
            return records
        finally:
            workbook.close()

    async def _encode_full(self, records: Iterable[Mapping[str, Any]]) -> bytes:
        xlsxwriter = self._load_xlsxwriter()
        materialised: list[Mapping[str, Any]] = list(records)
        if self._column_names is not None:
            columns = self._column_names
        elif materialised:
            columns = tuple(materialised[0].keys())
        else:
            columns = ()
        buf = io.BytesIO()
        workbook = xlsxwriter.Workbook(buf, {"in_memory": True})
        try:
            worksheet = workbook.add_worksheet(self._sheet_name)
            row_offset = 0
            if self._has_header:
                for col_index, column in enumerate(columns):
                    worksheet.write(0, col_index, column)
                row_offset = 1
            for row_index, record in enumerate(materialised):
                for col_index, column in enumerate(columns):
                    value = record.get(column)
                    worksheet.write(row_index + row_offset, col_index, value)
        finally:
            workbook.close()
        return buf.getvalue()

    @staticmethod
    def _load_openpyxl() -> Any:
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError(
                "XlsxFormat requires openpyxl. Install with `pip install pirn[xlsx]`."
            ) from exc
        return openpyxl

    @staticmethod
    def _load_xlsxwriter() -> Any:
        try:
            import xlsxwriter
        except ImportError as exc:
            raise ImportError(
                "XlsxFormat requires xlsxwriter. Install with `pip install pirn[xlsx]`."
            ) from exc
        return xlsxwriter
